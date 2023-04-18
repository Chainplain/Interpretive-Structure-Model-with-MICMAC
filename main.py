import numpy as np
from ISM_simple import Interpretive_structure_model as ISM
from openpyxl import load_workbook


if __name__ == "__main__":
    file_name = input("Please input the .xlsx filename which contains data:")

    ad_mat_excel=load_workbook(file_name+".xlsx")
    print("Loading" + ad_mat_excel.sheetnames[0])
    table = ad_mat_excel.get_sheet_by_name( ad_mat_excel.sheetnames[0])
    max_rows = table.max_row
    max_cols = table.max_column 
    print("max_row:",max_rows,"max_column:",max_cols)
    print("type of table:",type(table))
    table_data = tuple(table)

    inputmat = np.zeros([max_rows-1,max_cols-1])
    for i in range(max_rows-1):
        for j in range(max_cols-1):
            inputmat[i,j] = table_data[i+1][j+1].value
    print("inputmat:",inputmat)
    

    ism_hehe = ISM(file_name, inputmat)
    out = ism_hehe.Get_re_mat()
    print(out)
    
    ism_hehe.Compute_reach_sets()
    ism_hehe.Compute_priori_sets()
    ism_hehe.Compute_level_sets()
    ism_hehe.Compute_MICMAC()

    ism_hehe.Img_show()

    
